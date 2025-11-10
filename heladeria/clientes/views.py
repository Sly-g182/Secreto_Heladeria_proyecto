from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.db.models import Q, Count, Sum, Max

from ventas.models import Venta
from clientes.models import Cliente
from .forms import ClienteUserCreationForm, EditarPerfilForm, CambiarPasswordForm

import openpyxl
from openpyxl.styles import Font


def register_view(request):
    if request.method == 'POST':
        form = ClienteUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario registrado correctamente. Ahora puedes iniciar sesión.")
            return redirect('clientes:login')
    else:
        form = ClienteUserCreationForm()
    return render(request, 'clientes/register.html', {'form': form})


def login_view(request):
    next_url = request.POST.get('next') or request.GET.get('next')
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if not next_url:
                next_url = 'clientes:inicio'
            messages.success(request, f"¡Bienvenido, {user.username}!")
            return redirect(next_url)
        else:
            messages.error(request, "Nombre de usuario o contraseña incorrectos.")
    else:
        form = AuthenticationForm()
    return render(request, 'clientes/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.info(request, "Has cerrado sesión correctamente.")
    return redirect('clientes:login')


def inicio_clientes(request):
    return render(request, 'inicio.html')


@login_required
def historial_pedidos(request):
    cliente = request.user.cliente
    ventas = Venta.objects.filter(cliente=cliente).order_by('-fecha_venta')
    return render(request, 'ventas/historial.html', {'pedidos': ventas})


@login_required
def editar_perfil(request):
    """Vista para editar los datos del perfil del usuario, incluyendo imagen de perfil."""
    cliente = Cliente.objects.get(user=request.user)

    if request.method == 'POST':
        form = EditarPerfilForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Tu perfil se ha actualizado correctamente.")
            return redirect('clientes:editar_perfil')
        else:
            messages.error(request, "Por favor corrige los errores del formulario.")
    else:
        form = EditarPerfilForm(instance=request.user)

    return render(request, 'clientes/editar.html', {'form': form, 'cliente': cliente})


@login_required
def cambiar_password(request):
    """Vista para cambiar la contraseña del usuario."""
    if request.method == 'POST':
        form = CambiarPasswordForm(request.POST)
        if form.is_valid():
            nueva_password = form.cleaned_data.get('nueva_password')
            user = request.user
            user.set_password(nueva_password)
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Tu contraseña ha sido cambiada correctamente.")
            return redirect('clientes:editar_perfil')
        else:
            messages.error(request, "Por favor corrige los errores del formulario.")
    else:
        form = CambiarPasswordForm()

    return render(request, 'clientes/cambiar_password.html', {'form': form})


# ------------------------------
#     REPORTE Y EXPORTACIÓN
# ------------------------------

@staff_member_required
@staff_member_required
def reporte_clientes(request):
    query = request.GET.get('q', '')

    clientes = Cliente.objects.select_related('user').annotate(
        total_ordenes=Count('venta'),
        monto_total_gastado=Sum('venta__total'),
        ultima_compra=Max('venta__fecha')
    )

    if query:
        clientes = clientes.filter(Q(user__username__icontains=query))

    clientes = clientes.order_by('-monto_total_gastado')

    context = {
        'datos_clientes': clientes,
        'now': timezone.now(),
    }
    return render(request, 'clientes/reporte_clientes.html', context)




@staff_member_required
def exportar_clientes_excel(request):
    """Genera un archivo Excel con todos los clientes registrados y sus métricas."""

    # Crear el workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Encabezados
    columnas = [
        "ID", "Usuario", "Correo", "RUT", "Teléfono",
        "Dirección", "Fecha de Registro", "Última Compra",
        "Total de Órdenes", "Monto Total Gastado"
    ]
    for col_num, columna in enumerate(columnas, 1):
        c = ws.cell(row=1, column=col_num)
        c.value = columna
        c.font = Font(bold=True)

    # Traer datos con agregaciones
    clientes = Cliente.objects.select_related('user').annotate(
        total_ordenes=Count('ventas'),
        monto_total_gastado=Sum('ventas__total'),
        ultima_compra=Max('ventas__fecha_venta')
    )

    # Llenar datos en la hoja
    for row_num, cliente in enumerate(clientes, 2):
        ws.cell(row=row_num, column=1, value=cliente.id)
        ws.cell(row=row_num, column=2, value=cliente.user.username)
        ws.cell(row=row_num, column=3, value=cliente.user.email)
        ws.cell(row=row_num, column=4, value=cliente.rut or "N/A")
        ws.cell(row=row_num, column=5, value=cliente.telefono or "N/A")
        ws.cell(row=row_num, column=6, value=cliente.direccion or "N/A")
        ws.cell(row=row_num, column=7, value=cliente.user.date_joined.strftime('%Y-%m-%d'))

        if cliente.ultima_compra:
            ws.cell(row=row_num, column=8, value=cliente.ultima_compra.strftime('%Y-%m-%d %H:%M'))
        else:
            ws.cell(row=row_num, column=8, value="N/A")

        ws.cell(row=row_num, column=9, value=cliente.total_ordenes or 0)
        ws.cell(row=row_num, column=10, value=float(cliente.monto_total_gastado or 0))

    # Configurar respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=\"reporte_clientes.xlsx\"'
    wb.save(response)
    return response


